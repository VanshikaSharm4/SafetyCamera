import cv2
import copy
import csv
import itertools
import os
import time
from collections import Counter, deque
from threading import Thread
import threading

import numpy as np
import mediapipe as mp
import openpyxl

from weapon_detection import WeaponDetector
from model.keypoint_classifier.keypoint_classifier import KeyPointClassifier
from model.point_history_classifier.point_history_classifier import PointHistoryClassifier
from telegram_alerts import send_telegram_alert
from utils import CvFpsCalc
# SafetyCam class is defined below

class SafetyCam:
    def __init__(self, camera_url=0, show_feed=False):
        self.camera_url = camera_url
        self.show_feed = show_feed
        self.running = False
        self.cap = None
        self.latest_frame = None

        self.weapon_detector = WeaponDetector()
        self.mp_hands = mp.solutions.hands
        self.hands = self.mp_hands.Hands(
            static_image_mode=False,
            max_num_hands=2,
            min_detection_confidence=0.7,
            min_tracking_confidence=0.5,
        )

        self.keypoint_classifier = KeyPointClassifier()
        self.point_history_classifier = PointHistoryClassifier()

        with open('model/keypoint_classifier/keypoint_classifier_label.csv', encoding='utf-8-sig') as f:
            self.keypoint_classifier_labels = [row[0] for row in csv.reader(f)]

        with open('model/point_history_classifier/point_history_classifier_label.csv', encoding='utf-8-sig') as f:
            self.point_history_classifier_labels = [row[0] for row in csv.reader(f)]

        self.history_length = 16
        self.point_history = deque(maxlen=self.history_length)
        self.finger_gesture_history = deque(maxlen=self.history_length)
        self.cvFpsCalc = CvFpsCalc(buffer_len=10)

        self.wb = openpyxl.load_workbook('data.xlsx')
        self.sheet = self.wb.active

    def start(self):
        if self.running:
            return
        self.running = True
        self.thread = Thread(target=self._run, daemon=True)
        self.thread.start()

    def stop(self):
        self.running = False
        if self.cap:
            self.cap.release()
        cv2.destroyAllWindows()

    def _run(self):
        self.cap = cv2.VideoCapture(self.camera_url)
        last_excel_save = time.time()
        excel_dirty = False
        while self.running:
            fps = self.cvFpsCalc.get()
            ret, image = self.cap.read()
            if not ret:
                break

            # Resize for speed and quality
            image = cv2.resize(image, (416, 416))

            # Weapon detection (already threads alert)
            image = self.weapon_detector.detect_and_save(image)
            image = cv2.flip(image, 1)
            debug_image = image.copy()

            image = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
            image.flags.writeable = False
            results = self.hands.process(image)
            image.flags.writeable = True

            if results.multi_hand_landmarks is not None:
                for hand_landmarks, handedness in zip(results.multi_hand_landmarks, results.multi_handedness):
                    brect = self._calc_bounding_rect(debug_image, hand_landmarks)
                    landmark_list = self._calc_landmark_list(debug_image, hand_landmarks)
                    pre_processed_landmarks = self._pre_process_landmark(landmark_list)
                    pre_processed_point_history = self._pre_process_point_history(debug_image)

                    hand_sign_id = self.keypoint_classifier(pre_processed_landmarks)

                    if hand_sign_id == 2:
                        filename = f'data/gesture_warning_{int(time.time())}.jpg'
                        cv2.imwrite(filename, debug_image)
                        self._send_alert_async('Warning gesture detected!', filename)
                        self.sheet.cell(row=1, column=1).value = 1
                        excel_dirty = True
                    elif hand_sign_id == 3:
                        filename = f'data/gesture_distress_{int(time.time())}.jpg'
                        cv2.imwrite(filename, debug_image)
                        self._send_alert_async('Distress gesture detected!', filename)
                        self.sheet.cell(row=1, column=2).value = 1
                        excel_dirty = True
                        if not os.path.exists('data'):
                            os.makedirs('data')
                        count = 0
                        while count != 140:
                            suc, frame = self.cap.read()
                            if count % 20 == 0:
                                cv2.imwrite(f'./data/frame{count}.jpg', frame)
                            count += 1

                    finger_gesture_id = 0
                    if len(pre_processed_point_history) == (self.history_length * 2):
                        finger_gesture_id = self.point_history_classifier(pre_processed_point_history)

                    self.finger_gesture_history.append(finger_gesture_id)
                    most_common_fg_id = Counter(self.finger_gesture_history).most_common()

                    debug_image = self._draw_annotations(debug_image, brect, landmark_list,
                                                        handedness, hand_sign_id,
                                                        most_common_fg_id[0][0])
            else:
                self.point_history.append([0, 0])

            debug_image = self._draw_point_history(debug_image)
            debug_image = self._draw_info(debug_image, fps)

            self.latest_frame = debug_image.copy()

            # Save to Excel in background if dirty and 2s passed
            if excel_dirty and (time.time() - last_excel_save > 2):
                self._save_excel_async()
                last_excel_save = time.time()
                excel_dirty = False

        self.cap.release()
        cv2.destroyAllWindows()

    def _send_alert_async(self, message, image_path=None):
        threading.Thread(target=send_telegram_alert, args=(message,), kwargs={'image_path': image_path}, daemon=True).start()

    def _save_excel_async(self):
        threading.Thread(target=self.wb.save, args=("data.xlsx",), daemon=True).start()

    # Helper methods copied from testt.py (calc, pre_process, draw, etc.)
    def _calc_bounding_rect(self, image, landmarks):
        image_width, image_height = image.shape[1], image.shape[0]
        landmark_array = np.array([[int(lm.x * image_width), int(lm.y * image_height)] for lm in landmarks.landmark])
        x, y, w, h = cv2.boundingRect(landmark_array)
        return [x, y, x + w, y + h]

    def _calc_landmark_list(self, image, landmarks):
        image_width, image_height = image.shape[1], image.shape[0]
        return [[int(lm.x * image_width), int(lm.y * image_height)] for lm in landmarks.landmark]

    def _pre_process_landmark(self, landmark_list):
        base_x, base_y = landmark_list[0]
        relative = [[x - base_x, y - base_y] for x, y in landmark_list]
        flattened = list(itertools.chain.from_iterable(relative))
        max_val = max(map(abs, flattened))
        return [v / max_val for v in flattened]

    def _pre_process_point_history(self, image):
        image_width, image_height = image.shape[1], image.shape[0]
        base_x, base_y = self.point_history[0] if self.point_history else (0, 0)
        normalized = [[(x - base_x) / image_width, (y - base_y) / image_height] for x, y in self.point_history]
        return list(itertools.chain.from_iterable(normalized))

    def _draw_annotations(self, image, brect, landmark_list, handedness, hand_sign_id, finger_gesture_id):
        cv2.rectangle(image, (brect[0], brect[1]), (brect[2], brect[3]), (0, 255, 0), 2)
        label = self.keypoint_classifier_labels[hand_sign_id]
        fg_label = self.point_history_classifier_labels[finger_gesture_id]
        cv2.putText(image, f"{handedness.classification[0].label}: {label}", (brect[0], brect[1] - 10),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 255, 255), 1)
        return image

    def _draw_point_history(self, image):
        for i, (x, y) in enumerate(self.point_history):
            if x and y:
                cv2.circle(image, (x, y), 1 + i // 2, (0, 255, 0), 2)
        return image

    def _draw_info(self, image, fps):
        return image
